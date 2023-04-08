from tkinter import *
#for sending dialogue message box
from tkinter import messagebox

import datetime
import os
import sys 
import pandas as pd
import openpyxl
#for mailing
import smtplib
from email.message import EmailMessage
import ssl
#for sending attachments
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

####################################################################################################

                                    #MAILING
    
####################################################################################################

#data segregation
def seggregation():
    global remail

    mdf=pd.read_csv('Admin_data.csv')
    mdf=mdf[mdf['A_Sheet']==s_1]

    sample=mdf['A_Sheet']
    sample=str(sample)
    s1=""
    flag=0
    startind=0
    for i in range(0,len(sample)):
        if(flag==0):
            if((sample[i]>='0' and sample[i]<='9') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='a' and sample[i]<='z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='A' and sample[i]<='Z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
        else:
            break
    sample=sample[i-1:]
    for i in sample:
        if(i=="\n" or i==' '):
            break
        s1+=i
    s1+=".xlsx"
    
    wb1=openpyxl.load_workbook(s1)
    name='data'
    sheet=wb1[name]
    r=0
    f=0
    mail_store=''
    for row in sheet.iter_rows(min_row=2, min_col=2, max_row=350, max_col=2):
        for c in row:
            r+=1
            if(c.value==rdata):
                s=sheet.cell(row=r+1,column=3)
                mail_store=s.value
                f=1
                break
        if f==1:
            break
    wb1.save(s1)
    remail=mail_store
#sending Email
def sendemail():
    e_sender='drplusdevops@gmail.com'
    seggregation()
    e_pass='tszburfhowxtjyyw'
    
    e_receiver=remail

    sub="Attendance report."
    body="Attendance captured for "+t_d+' on '+str(datetime.datetime.now())

    em=EmailMessage()
    em['From']=e_sender
    em['To']=e_receiver
    em['Subject']=sub
    em.set_content(body)

    context=ssl.create_default_context()
    with smtplib.SMTP_SSL('smtp.gmail.com',465,context=context) as smt:
        smt.login(e_sender,e_pass)
        smt.sendmail(e_sender,e_receiver,em.as_string())




        

#*******************************************************************************************8
#sending mail to colleges
def sendemailclg(cmail,file):
    port=465
    smtp_server='smtp.gmail.com'
    
    e_sender='drplusdevops@gmail.com'
    e_pass='tszburfhowxtjyyw'
    e_receiver=cmail
          
    sub="Attendance report."
    body="Attendance report on "+str(datetime.datetime.now())

    aem=MIMEMultipart()
    aem['From']=e_sender
    aem['To']=e_receiver
    aem['Subject']=sub

    aem.attach(MIMEText(body,'plain'))

    filename=file

    attachment=open(filename,'rb')
    attachment_package=MIMEBase('application','octet-stream')
    attachment_package.set_payload((attachment).read())
    encoders.encode_base64(attachment_package)
    attachment_package.add_header('Content-Disposition',"attachment; filename= "+filename)
    aem.attach(attachment_package)

    context=ssl.create_default_context()
    with smtplib.SMTP_SSL('smtp.gmail.com',465,context=context) as smt:
        smt.login(e_sender,e_pass)
        smt.sendmail(e_sender,e_receiver,aem.as_string())

###################################################################################################

                                    #Mailing END
    
####################################################################################################
    
#***************************************************************************************************************************
#***************************************************************************************************************************
        
###################################################################################################

                                    #TRAINER
    
####################################################################################################
    
#scanning into csv file
df=pd.DataFrame(columns=['Roll_No'])
def writecsv():
    global df
    global rdata
    
    rdata=data.get()
    rdata=rdata.upper()

    lr=[rdata]

    new_row=pd.DataFrame({'Roll_No':rdata}, index=[0])
    df = pd.concat([new_row,df.loc[:]]).reset_index(drop=True)

    save_df=pd.read_csv('Admin_data.csv')
    save_df=save_df[save_df['A_Sheet']==s_1]

    sample=save_df['A_Sheet']
    sample=str(sample)
    s1=""
    flag=0
    startind=0
    for i in range(0,len(sample)):
        if(flag==0):
            if((sample[i]>='0' and sample[i]<='9') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='a' and sample[i]<='z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='A' and sample[i]<='Z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
        else:
            break
    sample=sample[i-1:]
    for i in sample:
        if(i=="\n" or i==' '):
            break
        s1+=i
    s1+=".xlsx"

    wb1=openpyxl.load_workbook(s1)
    sheet = wb1[rdata]
    r=0
    f=0
    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=350, max_col=1):
        for c in row:
            r+=1
            if(c.value==None):
                f=1
                break
        if f==1:
            break
    r+=1
    cell1=sheet.cell(row=r,column=1)
    cell2=sheet.cell(row=r,column=2)

    datetime_obj=datetime.datetime.now()
    date=datetime_obj.date()

    cell1.value=str(date)
    cell2.value='Present'
    
    wb1.save(s1)    

    with pd.ExcelWriter(s1, mode="a", engine="openpyxl",if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name="Attendance")
    
    #sending email
    sendemail()
   
#scan() funtion
def scan():
    ldata=data.get()
    ldata=ldata.upper()
    if(ldata==''):
        lab1.configure(text="Please enter a Roll No")
    elif ldata[2]+ldata[3] == "A9" or ldata[2]+ldata[3] == "P3" or ldata[2]+ldata[3] == "MH" :
        lab1.configure(text=ldata+" done!")
        writecsv()
        reset()
    else:
        lab1.configure(text="Please enter a valid Roll No!")


#reset()
def reset():
    den.delete(0,END)

#tabdel
def td():

    save_df=pd.read_csv('Admin_data.csv')
    save_df=save_df[save_df['A_Sheet']==s_1]

    sample=save_df['A_Sheet']
    sample=str(sample)
    s1=""
    flag=0
    startind=0
    for i in range(0,len(sample)):
        if(flag==0):
            if((sample[i]>='0' and sample[i]<='9') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='a' and sample[i]<='z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='A' and sample[i]<='Z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
        else:
            break
    sample=sample[i-1:]
    for i in sample:
        if(i=="\n" or i==' '):
            break
        s1+=i
    s1+=".xlsx"

    wb1=openpyxl.load_workbook(s1)
    name='Attendance'
    li=wb1.sheetnames
    if name in li:
        del wb1['Attendance']
    wb1.save(s1)

    li=[]
    wb1=openpyxl.load_workbook(s1)
    name='data'
    f=0
    sheet=wb1[name]
    for row in sheet.iter_rows(min_row=2, min_col=2, max_row=350, max_col=2):
        for c in row:
            if(c.value==None):
                f=1
                break
            li.append(c.value)
        if f==1:
            break
    wb1.save(s1)

    for i in li:
        name=i
        if i not in df['Roll_No'].unique():
            wb1=openpyxl.load_workbook(s1)
            sheet=wb1[name]
            r=0
            f=0
            for row in sheet.iter_rows(min_row=2, min_col=1, max_row=350, max_col=1):
                for c in row:
                    r+=1
                    if(c.value==None):
                        f=1
                        break
                if f==1:
                    break
            r+=1
            cell1=sheet.cell(row=r,column=1)
            cell2=sheet.cell(row=r,column=2)
            datetime_obj=datetime.datetime.now()
            date=datetime_obj.date()

            cell1.value=str(date)
            cell2.value='Absent'
            
            wb1.save(s1)
    
    with pd.ExcelWriter(s1, mode="a", engine="openpyxl",if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name="Attendance")
    
    tab.destroy()
    trainer_properties(user_t,t_d,s_1)
    
#Scanning tab
def scantab():

    global tab
    global data
    global den
    global lab1
    
    tab=Tk()
    tab.title("Attendance")
    tab.geometry("1920x1080")
    
    
    sc1=Frame(tab,bg="#e1f0db",width=550,height=300,bd=5).place(relx=0.5,rely=0.4,anchor=CENTER)
    sltitle=Label(text="Attendance",font=("Gabriola",30,'bold'),bg="#e1f0db",bd=2).place(relx=0.5,rely=0.3,anchor=CENTER)

    Label(tab,text="Scan your ID: ",font=("Sitka Small Semibold",15),bg="#72bd20").place(relx=0.5,rely=0.4,anchor=CENTER)
    
    data=StringVar()
    
    den=Entry(tab,textvariable=data,width=22,bd=2,font=("ariel FB",10))
    den.place(relx=0.5,rely=0.5,anchor=CENTER)

    lab1=Label(tab,font=("Comic Sans MS",12),bg="#e1f0db",fg="#0e6473")
    
    Button(tab,text="Done",font=("Sitka Small Semibold",5),height="2",width=15,bd=1,command=scan).place(relx=0.5,rely=0.6,anchor=CENTER)
    lab1.place(relx=0.5,rely=0.55,anchor=CENTER)

    Button(tab,text="Logout",font=("ariel",10),height=1,width=15,bd=1,command=td).place(relx=0.5,rely=0.75,anchor=CENTER)
    tab.mainloop()


#Scanning tab event

def scantabev():

    global tabev
    global datae
    global dene
    global labe1
    
    tabev=Tk()
    tabev.title("Attendance")
    tabev.geometry("1920x1080")
    
    
    sc1=Frame(tabev,bg="#e1f0db",width=550,height=300,bd=5).place(relx=0.5,rely=0.4,anchor=CENTER)
    sltitle=Label(text="Attendance",font=("Gabriola",30,'bold'),bg="#e1f0db",bd=2).place(relx=0.5,rely=0.3,anchor=CENTER)

    Label(tabev,text="Scan your ID: ",font=("Sitka Small Semibold",15),bg="#72bd20").place(relx=0.5,rely=0.4,anchor=CENTER)
    
    datae=StringVar()
    
    dene=Entry(tabev,textvariable=datae,width=22,bd=2,font=("ariel FB",10))
    dene.place(relx=0.5,rely=0.5,anchor=CENTER)

    labe1=Label(tabev,font=("Comic Sans MS",12),bg="#e1f0db",fg="#0e6473")
    
    Button(tabev,text="Done",font=("Sitka Small Semibold",5),height="2",width=15,bd=1,command=scanev).place(relx=0.5,rely=0.6,anchor=CENTER)
    labe1.place(relx=0.5,rely=0.55,anchor=CENTER)

    Button(tabev,text="Logout",font=("ariel",10),height=1,width=15,bd=1,command=tde).place(relx=0.5,rely=0.75,anchor=CENTER)
    tabev.mainloop()

#scan event
def scanev():
    ldata=datae.get()
    ldata=ldata.upper()
    if(ldata==''):
        labe1.configure(text="Please enter a Roll No")
    elif ldata[2]+ldata[3] == "A9" or ldata[2]+ldata[3] == "P3" or ldata[2]+ldata[3] == "MH" :
        labe1.configure(text=ldata+" done!")
        writeev()
    else:
        labe1.configure(text="Please enter a valid Roll No!")

#writing attendance 
edf=pd.DataFrame(columns=['Roll_No'])
def writeev():
    global edf
    
    t_rno=datae.get()
    t_rno=t_rno.upper()
    
    a_df=pd.read_csv('Admin_data.csv')
    ta_df=a_df[a_df['A_Sheet']==s_1]

    sample=ta_df['A_Sheet']
    sample=str(sample)
    s1=""
    flag=0
    startind=0
    for i in range(0,len(sample)):
        if(flag==0):
            if((sample[i]>='0' and sample[i]<='9') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='a' and sample[i]<='z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='A' and sample[i]<='Z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
        else:
            break
    sample=sample[i-1:]
    for i in sample:
        if(i=="\n" or i==' '):
            break
        s1+=i
    s1+=".xlsx"

    wb1=openpyxl.load_workbook(s1)
    name='data'
    sheet=wb1[name]
    r=0
    f=0
    for row in sheet.iter_rows(min_row=2, min_col=2, max_row=350, max_col=2):
        for c in row:
            r+=1
            if(c.value==None):
                f=1
                break
        if f==1:
            break
    r+=1
    cell2=sheet.cell(row=r,column=2)
    cell2.value=t_rno
    
    wb1.save(s1)
    dene.delete(0,END)

    new_row=pd.DataFrame({'Roll_No':t_rno}, index=[0])
    edf = pd.concat([new_row,edf.loc[:]]).reset_index(drop=True)

    with pd.ExcelWriter(s1, mode="a", engine="openpyxl",if_sheet_exists='overlay') as writer:
        edf.to_excel(writer, sheet_name="Attendance")


def tde():
    save_df=pd.read_csv('Admin_data.csv')
    save_df=save_df[save_df['A_Sheet']==s_1]

    sample=save_df['A_Sheet']
    sample=str(sample)
    s1=""
    flag=0
    startind=0
    for i in range(0,len(sample)):
        if(flag==0):
            if((sample[i]>='0' and sample[i]<='9') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='a' and sample[i]<='z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='A' and sample[i]<='Z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
        else:
            break
    sample=sample[i-1:]
    for i in sample:
        if(i=="\n" or i==' '):
            break
        s1+=i
    s1+=".xlsx"
    
    with pd.ExcelWriter(s1, mode="a", engine="openpyxl",if_sheet_exists='overlay') as writer:
        edf.to_excel(writer, sheet_name="Attendance")
    
    tabev.destroy()
    trainer_properties(user_t,t_d,s_1)

    
def login():
    check_id=id_.get()
    user=username.get()
    pas=password.get()

    tdf=pd.read_csv('Admin_data.csv')
    if(int(check_id) in tdf['ID'].unique()):
        stdf=tdf[ tdf['ID']==int(check_id) ]
        ############### domain
        sample=stdf['domain']
        sample=str(sample)
        s=""
        flag=0
        startind=0
        for i in range(0,len(sample)):
            if(flag==0):
                if((sample[i]>='0' and sample[i]<='9') and (sample[i+1]!=' ')):
                    startind=i
                    flag=1
                
                elif((sample[i]>='a' and sample[i]<='z') and (sample[i+1]!=' ')):
                    startind=i
                    flag=1
                
                elif((sample[i]>='A' and sample[i]<='Z') and (sample[i+1]!=' ')):
                    startind=i
                    flag=1
            else:
                break
        sample=sample[i-1:]
        for i in sample:
            if(i=="\n" or i==' '):
                break
            s+=i
        ######## Attendsnce sheet
        sample=stdf['A_Sheet']
        sample=str(sample)
        s1=""
        flag=0
        startind=0
        for i in range(0,len(sample)):
            if(flag==0):
                if((sample[i]>='0' and sample[i]<='9') and (sample[i+1]!=' ')):
                    startind=i
                    flag=1
                
                elif((sample[i]>='a' and sample[i]<='z') and (sample[i+1]!=' ')):
                    startind=i
                    flag=1
                
                elif((sample[i]>='A' and sample[i]<='Z') and (sample[i+1]!=' ')):
                    startind=i
                    flag=1
            else:
                break
        sample=sample[i-1:]
        for i in sample:
            if(i=="\n" or i==' '):
                break
            s1+=i
        
        if(int(pas) in stdf['password'].unique()):
            root.destroy()
            trainer_properties(user,s,s1)
        else:
            messagebox.showinfo("Invalid","Please enter correct password!")
        
    elif user=="" or pas=="":
        messagebox.showinfo("Invalid","Username and Password can't be empty!")
    elif user not in tdf['username'].unique():
        messagebox.showinfo("Invalid","Please enter a valid username!")
    else:
        messagebox.showinfo("Password Error","Please enter correct details!")

#trainer properties
def trainer_properties(usert,td,s1):
    global tp
    global user_t
    global t_d
    global s_1

    user_t=usert
    t_d=td
    s_1=s1

    tp=Tk()
    tp.title("Trainer")
    tp.geometry("1920x1080")
    tp.configure(bg="#ffffff")
 
    bc1=Frame(tp,bg="#fad311",width=650,height=450).place(relx=0.5,rely=0.45,anchor=CENTER)
    bc2=Frame(tp,bg="#72bd20",width=600,height=400).place(relx=0.5,rely=0.45,anchor=CENTER)

    img=PhotoImage(file="logo_header.png"
                   )
    img_label=Label(tp,image=img)
    img_label.grid(row=0,column=0)
    img_label.image=img

    Label(tp,text="Hello "+user_t+"!",height=0,font=("Gabriola",20,'bold'),bg="#72bd20").place(relx=0.5,rely=0.3,anchor=CENTER)
    
    Button(tp,text="View Trainees",height="1",width=17,bd=1,command=call_trainer_users).place(relx=0.5,rely=0.35,anchor=CENTER)
    Button(tp,text="Change Password",height="1",width=17,bd=1,command=call_trainer_change_pass).place(relx=0.5,rely=0.4,anchor=CENTER)
    Button(tp,text="Attendance",height="1",width=17,bd=1,command=call_scantab).place(relx=0.5,rely=0.45,anchor=CENTER)
    Button(tp,text="Attendance Tracker",height="1",width=17,bd=1,command=call_attendance_tracker).place(relx=0.5,rely=0.5,anchor=CENTER)
    Button(tp,text="Add Trainee",height="1",width=17,bd=1,command=call_attendance_sheet).place(relx=0.5,rely=0.55,anchor=CENTER)
    Button(tp,text="Delete Trainee",height="1",width=17,bd=1,command=call_delete_sheet).place(relx=0.5,rely=0.6,anchor=CENTER)
    Button(tp,text="Scan (Event)",height="1",width=17,bd=1,command=call_scantabev).place(relx=0.5,rely=0.65,anchor=CENTER)
    
    Button(tp,text="Exit",height="2",width=20,bd=1,command=call_tlogin).place(relx=0.5,rely=0.7,anchor=CENTER)

    tp.mainloop()

#calling traines view
def call_trainer_users():
    tp.destroy()
    trainer_users()

def trainer_users():
    global tu
    
    tu_df=pd.read_csv('Admin_data.csv')
    tu_df=tu_df[tu_df['A_Sheet']==s_1]

    sample=tu_df['A_Sheet']
    sample=str(sample)
    s1=""
    flag=0
    startind=0
    for i in range(0,len(sample)):
        if(flag==0):
            if((sample[i]>='0' and sample[i]<='9') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='a' and sample[i]<='z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='A' and sample[i]<='Z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
        else:
            break
    sample=sample[i-1:]
    for i in sample:
        if(i=="\n" or i==' '):
            break
        s1+=i
    s1+=".xlsx"

    wb1=openpyxl.load_workbook(s1)
    name='data'
    sheet=wb1[name]
    check=sheet.cell(row=1,column=3)
    wb1.save(s1)
    if(check.value==None):
        li1=[]
        wb1=openpyxl.load_workbook(s1)
        name='data'
        sheet=wb1[name]
        r=0
        f=0
        for row in sheet.iter_rows(min_row=2, min_col=2, max_row=350, max_col=2):
            for c in row:
                r+=1
                if(c.value==None):
                    f=1
                    break
                li1.append(c.value)
            if f==1:
                break
        wb1.save(s1)
        tudf=pd.DataFrame(columns=['ROll_No'])
        tudf['ROll_No']=li1
    else:
        li1=[]
        wb1=openpyxl.load_workbook(s1)
        name='data'
        sheet=wb1[name]
        r=0
        f=0
        for row in sheet.iter_rows(min_row=2, min_col=1, max_row=350, max_col=1):
            for c in row:
                r+=1
                if(c.value==None):
                    f=1
                    break
                li1.append(c.value)
            if f==1:
                break
        wb1.save(s1)

        li2=[]
        wb1=openpyxl.load_workbook(s1)
        name='data'
        sheet=wb1[name]
        r=0
        f=0
        for row in sheet.iter_rows(min_row=2, min_col=2, max_row=350, max_col=2):
            for c in row:
                r+=1
                if(c.value==None):
                    f=1
                    break
                li2.append(c.value)
            if f==1:
                break
        wb1.save(s1)

        li3=[]
        wb1=openpyxl.load_workbook(s1)
        name='data'
        sheet=wb1[name]
        r=0
        f=0
        for row in sheet.iter_rows(min_row=2, min_col=3, max_row=350, max_col=3):
            for c in row:
                r+=1
                if(c.value==None):
                    f=1
                    break
                li3.append(c.value)
            if f==1:
                break
        wb1.save(s1)

        tudf=pd.DataFrame(columns=['Index','ROll_No','Email'])
        tudf['Index']=li1
        tudf['ROll_No']=li2
        tudf['Email']=li3

    tu=Tk()
    tu.geometry('1920x1080')
    txt=Text(tu) 
    txt.pack() 

    class PrintToTXT(object): 
        def write(self, s): 
            txt.insert(END, s)

    sys.stdout = PrintToTXT() 
    print('Trainees Data') 
    print(tudf)
    
    Button(tu,text="Edit Data",height="1",width=20,bd=1,command=call_tu_edit).place(relx=0.5,rely=0.6,anchor=CENTER)
    Button(tu,text="Back",height="1",width=20,bd=1,command=call_tu_back).place(relx=0.5,rely=0.7,anchor=CENTER)

#editing the trainees data
def call_tu_edit():
    tu.destroy()
    tu_edit_display()

def tu_edit_display():
    global tue
    global tu_t
    global tu_t1
    global tu_tt
    global tu_t2

    tue=Tk()
    tue.title("Trainer")
    tue.geometry("1920x1080")
    tue.configure(bg="#ffffff")
 
    bc1=Frame(tue,bg="#fad311",width=650,height=450).place(relx=0.5,rely=0.45,anchor=CENTER)
    bc2=Frame(tue,bg="#72bd20",width=600,height=400).place(relx=0.5,rely=0.45,anchor=CENTER)

    img=PhotoImage(file="logo_header.png")
    img_label=Label(tue,image=img)
    img_label.grid(row=0,column=0)
    img_label.image=img

    Label(tue,text="Enter the Trainee Details to edit:",font=("Sitka Small Semibold",15),bg="#72bd20").place(relx=0.5,rely=0.35,anchor=CENTER)
    Label(tue,text="Roll No:",font=("Sitka Small Semibold",15),bg="#72bd20").place(relx=0.4,rely=0.45,anchor=CENTER)
    Label(tue,text="Email:",font=("Sitka Small Semibold",15),bg="#72bd20").place(relx=0.4,rely=0.5,anchor=CENTER)

    tu_t=StringVar()
    tu_tt=StringVar()
    
    tu_t1=Entry(tue,textvariable=tu_t,width=22,bd=2,font=("Californian FB",10))
    tu_t1.place(relx=0.5,rely=0.45,anchor=CENTER)
    tu_t2=Entry(tue,textvariable=tu_tt,width=22,bd=2,font=("Californian FB",10))
    tu_t2.place(relx=0.5,rely=0.5,anchor=CENTER)

    Button(tue,text="edit",height="1",width=10,bd=1,command=tu_edit).place(relx=0.5,rely=0.6,anchor=CENTER)
    Button(tue,text="Back",height="1",width=10,bd=1,command=tu_edit_back).place(relx=0.5,rely=0.7,anchor=CENTER)

    tue.mainloop()

def tu_edit_back():
    tue.destroy()
    trainer_properties(user_t,t_d,s_1)
    
def tu_edit():
    tue_df=pd.read_csv('Admin_data.csv')
    tue_df=tue_df[tue_df['A_Sheet']==s_1]

    sample=tue_df['A_Sheet']
    sample=str(sample)
    s1=""
    flag=0
    startind=0
    for i in range(0,len(sample)):
        if(flag==0):
            if((sample[i]>='0' and sample[i]<='9') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='a' and sample[i]<='z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='A' and sample[i]<='Z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
        else:
            break
    sample=sample[i-1:]
    for i in sample:
        if(i=="\n" or i==' '):
            break
        s1+=i
    s1+=".xlsx"

    roll=tu_t.get()
    em=tu_tt.get()
    if(roll=='' or em==''):
        messagebox.showinfo("Invalid","Input can't be empty!")
    else:
        roll=roll.upper()
        wb1=openpyxl.load_workbook(s1)
        name='data'
        sheet=wb1[name]
        r=0
        f=0
        for row in sheet.iter_rows(min_row=2, min_col=2, max_row=350, max_col=2):
            for c in row:
                r+=1
                if(c.value==roll):
                    f=1
                    break
                if(c.value==None):
                    f=1
                    break
            if f==1:
                break
        r+=1
        cell1=sheet.cell(row=r,column=2)
        cell2=sheet.cell(row=r,column=3)
        cell1.value=roll
        cell2.value=em
        
        wb1.save(s1)
        Label(tue,text="Done",font=("Comic Sans MS",10),bg="#72bd20").place(relx=0.5,rely=0.55,anchor=CENTER)


def call_tu_back():
    tu.destroy()
    trainer_properties(user_t,t_d,s_1)

#calling trainee deleting tab
def call_delete_sheet():
    tp.destroy()
    delete_trainee()

#adding a trainee
def delete_trainee():
    global d_t
    global del_t
    global del_t1

    d_t=Tk()
    d_t.title("Trainer")
    d_t.geometry("1920x1080")
    d_t.configure(bg="#ffffff")

    bc1=Frame(d_t,bg="#fad311",width=650,height=450).place(relx=0.5,rely=0.45,anchor=CENTER)
    bc2=Frame(d_t,bg="#72bd20",width=600,height=400).place(relx=0.5,rely=0.45,anchor=CENTER)

    img=PhotoImage(file="logo_header.png")
    img_label=Label(d_t,image=img)
    img_label.grid(row=0,column=0)
    img_label.image=img

    Label(d_t,text="Enter the Trainee Roll No to delete:",font=("Sitka Small Semibold",15),bg="#72bd20").place(relx=0.5,rely=0.4,anchor=CENTER)

    del_t=StringVar()

    del_t1=Entry(d_t,textvariable=del_t,width=22,bd=2,font=("Californian FB",10))
    del_t1.place(relx=0.5,rely=0.5,anchor=CENTER)

    Button(d_t,text="delete",height="1",width=10,bd=1,command=delete_sheet).place(relx=0.5,rely=0.6,anchor=CENTER)
    Button(d_t,text="Return to Dashboard",height="1",width=20,bd=1,command=call_dash_del).place(relx=0.5,rely=0.7,anchor=CENTER)

    d_t.mainloop()



def call_dash_del():
    d_t.destroy()
    trainer_properties(user_t,t_d,s_1)
    
#attendace sheet function
def delete_sheet():

    t_rno=del_t.get()
    t_rno=t_rno.upper()
    
    d_df=pd.read_csv('Admin_data.csv')
    td_df=d_df[d_df['A_Sheet']==s_1]

    sample=td_df['A_Sheet']
    sample=str(sample)
    s1=""
    flag=0
    startind=0
    for i in range(0,len(sample)):
        if(flag==0):
            if((sample[i]>='0' and sample[i]<='9') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='a' and sample[i]<='z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='A' and sample[i]<='Z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
        else:
            break
    sample=sample[i-1:]
    for i in sample:
        if(i=="\n" or i==' '):
            break
        s1+=i
    s1+=".xlsx"
    
    workbook=openpyxl.load_workbook(s1)
    del workbook[t_rno]
    workbook.save(s1)

    Label(d_t,text=t_rno+" deleted.",font=("Comic Sans MS",10),bg="#72bd20").place(relx=0.5,rely=0.55,anchor=CENTER)

    del_t1.delete(0,END)

    
#calling trainee adding tab
def call_attendance_sheet():
    tp.destroy()
    add_trainee()

#adding a trainee
def add_trainee():
    global a_t
    global add_t
    global add_m
    global add_m1
    global add_t1

    a_t=Tk()
    a_t.title("Trainer")
    a_t.geometry("1920x1080")
    a_t.configure(bg="#ffffff")

    bc1=Frame(a_t,bg="#fad311",width=650,height=450).place(relx=0.5,rely=0.45,anchor=CENTER)
    bc2=Frame(a_t,bg="#72bd20",width=600,height=400).place(relx=0.5,rely=0.45,anchor=CENTER)

    img=PhotoImage(file="logo_header.png")
    img_label=Label(a_t,image=img)
    img_label.grid(row=0,column=0)
    img_label.image=img

    Label(a_t,text="Enter the Trainee Roll No:",font=("Sitka Small Semibold",15),bg="#72bd20").place(relx=0.5,rely=0.4,anchor=CENTER)
    Label(a_t,text="Enter the Trainee Mail ID:",font=("Sitka Small Semibold",15),bg="#72bd20").place(relx=0.5,rely=0.5,anchor=CENTER)

    add_t=StringVar()
    add_m=StringVar()

    add_t1=Entry(a_t,textvariable=add_t,width=22,bd=2,font=("Californian FB",10))
    add_t1.place(relx=0.5,rely=0.45,anchor=CENTER)
    add_m1=Entry(a_t,textvariable=add_m,width=22,bd=2,font=("Californian FB",10))
    add_m1.place(relx=0.5,rely=0.55,anchor=CENTER)

    Button(a_t,text="add",height="1",width=10,bd=1,command=attendance_sheet).place(relx=0.5,rely=0.65,anchor=CENTER)
    Button(a_t,text="Return to Dashboard",height="1",width=20,bd=1,command=call_dash).place(relx=0.5,rely=0.7,anchor=CENTER)

    a_t.mainloop()

#calling user dashboard
def call_dash():
    a_t.destroy()
    trainer_properties(user_t,t_d,s_1)

#attendace sheet function
def attendance_sheet():

    t_rno=add_t.get()
    t_rno=t_rno.upper()

    t_mail=add_m.get()
    t_mail=t_mail.upper()
    
    a_df=pd.read_csv('Admin_data.csv')
    ta_df=a_df[a_df['A_Sheet']==s_1]

    sample=ta_df['A_Sheet']
    sample=str(sample)
    s1=""
    flag=0
    startind=0
    for i in range(0,len(sample)):
        if(flag==0):
            if((sample[i]>='0' and sample[i]<='9') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='a' and sample[i]<='z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='A' and sample[i]<='Z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
        else:
            break
    sample=sample[i-1:]
    for i in sample:
        if(i=="\n" or i==' '):
            break
        s1+=i
    s1+=".xlsx"

    wb1=openpyxl.load_workbook(s1)
    name='data'
    sheet=wb1[name]
    r=0
    f=0
    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=350, max_col=1):
        for c in row:
            r+=1
            if(c.value==None):
                f=1
                break
        if f==1:
            break
    r+=1
    cell1=sheet.cell(row=r,column=1)
    cell2=sheet.cell(row=r,column=2)
    cell3=sheet.cell(row=r,column=3)
    cell1.value=r-1
    cell2.value=t_rno
    cell3.value=t_mail
    
    wb1.save(s1)

    a_data=pd.DataFrame(columns=['date','Pre/Abs'])
    a_data.set_index("date", inplace = True)
    
    with pd.ExcelWriter(s1, mode="a", engine="openpyxl",if_sheet_exists='overlay') as writer:
        a_data.to_excel(writer, sheet_name=t_rno)

    Label(a_t,text=t_rno+" added.",font=("Comic Sans MS",10),bg="#72bd20").place(relx=0.5,rely=0.6,anchor=CENTER)

    add_t1.delete(0,END)
    add_m1.delete(0,END)

#calling chaning trainer password
def call_trainer_change_pass():
    tp.destroy()
    trainer_change_pass()

#changing trainer password
def trainer_change_pass():
    global cpass_t
    global tc_pass1
    global tc_pass2
    global t_cpass1
    global t_cpass2

    cpass_t=Tk()
    cpass_t.geometry("1920x1080")
    cpass_t.configure(bg="#ffffff")
    
    bc1=Frame(cpass_t,bg="#fad311",width=650,height=450).place(relx=0.5,rely=0.45,anchor=CENTER)
    bc2=Frame(cpass_t,bg="#72bd20",width=600,height=400).place(relx=0.5,rely=0.45,anchor=CENTER)

    Label(cpass_t,text="Please enter new Password",height=0,font=("Gabriola",20,'bold'),bg="#72bd20").place(relx=0.5,rely=0.3,anchor=CENTER)

    Label(cpass_t,text="Password: ",font=("Sitka Small Semibold",12),bg="#72bd20").place(relx=0.46,rely=0.42,anchor=CENTER)
    Label(cpass_t,text="Re-enter Password: ",font=("Sitka Small Semibold",12),bg="#72bd20").place(relx=0.43,rely=0.5,anchor=CENTER)

    tc_pass1=StringVar()
    tc_pass2=StringVar()
    
    t_cpass1=Entry(cpass_t,textvariable=tc_pass1,width=22,bd=2,font=("Californian FB",10))
    t_cpass1.place(relx=0.55,rely=0.42,anchor=CENTER)
    t_cpass2=Entry(cpass_t,textvariable=tc_pass2,width=22,bd=2,font=("Californian FB",10))
    t_cpass2.place(relx=0.55,rely=0.5,anchor=CENTER)

    img=PhotoImage(file="logo_header.png")
    img_label=Label(cpass_t,image=img)
    img_label.grid(row=0,column=0)
    img_label.image=img

    Button(cpass_t,text="change",height="1",width=10,bd=1,command=call_changet).place(relx=0.5,rely=0.6,anchor=CENTER)
    Button(cpass_t,text="Return Home screen",height="1",width=20,bd=1,command=call_tloginch).place(relx=0.5,rely=0.7,anchor=CENTER)

    cpass_t.mainloop()

#calling chcnaget
def call_changet():
    p1=tc_pass1.get()
    p2=tc_pass2.get()


    if p1=="" or p2=="":
        messagebox.showinfo("Invalid","password can't be empty!")
        t_cpass1.delete(0,END)
        t_cpass2.delete(0,END)
        
    elif(p1 == p2):
        df=pd.read_csv('Admin_data.csv')
        df.set_index("ID", inplace=True)
        idx=df.index[df['username']==user_t]
        df.loc[idx,'password']=p1
        df.to_csv('Admin_data.csv')
        
        lab1=Label(cpass_t,text="Password changed",font=("Comic Sans MS",12),bg="#72bd20",fg="#0e6473")
        lab1.place(relx=0.5,rely=0.55,anchor=CENTER)
        t_cpass1.delete(0,END)
        t_cpass2.delete(0,END)
        
    elif p1!=p2:
        messagebox.showinfo("Invalid","passwords didn't matched!")
        t_cpass1.delete(0,END)
        t_cpass2.delete(0,END)
    
#calling trainer home from change pass
def call_tloginch():
    cpass_t.destroy()
    trainerlogintab()

#calling scanning tab
def call_scantab():
    tp.destroy()
    scantab()

#calling scanning tab
def call_scantabev():
    tp.destroy()
    scantabev()
#*****************************************************************************************************************

def call_attendance_tracker():
    tp.destroy()
    attendance_tracker()
#attendance tracker
def attendance_tracker():
    global att_t
    global rno
    global rno1
    
    att_t=Tk()
    att_t.geometry("1920x1080")
    att_t.configure(bg="#ffffff")
    
    bc1=Frame(att_t,bg="#fad311",width=650,height=450).place(relx=0.5,rely=0.45,anchor=CENTER)
    bc2=Frame(att_t,bg="#72bd20",width=600,height=400).place(relx=0.5,rely=0.45,anchor=CENTER)

    Label(att_t,text="Enter the Roll No:",height=0,font=("Gabriola",20,'bold'),bg="#72bd20").place(relx=0.5,rely=0.3,anchor=CENTER)
    
    rno=StringVar()
    
    rno1=Entry(att_t,textvariable=rno,width=22,bd=2,font=("Californian FB",10))
    rno1.place(relx=0.5,rely=0.42,anchor=CENTER)
    
    img=PhotoImage(file="logo_header.png")
    img_label=Label(att_t,image=img)
    img_label.grid(row=0,column=0)
    img_label.image=img

    Button(att_t,text="Fetch",height="1",width=10,bd=1,command=fetch_student_calender).place(relx=0.45,rely=0.62,anchor=CENTER)
    Button(att_t,text="Back",height="1",width=10,bd=1,command=at_home).place(relx=0.55,rely=0.62,anchor=CENTER)           

    att_t.mainloop()

#returning to trainer properties from attendance tracker
def at_home():
    att_t.destroy()
    trainer_properties(user_t,t_d,s_1)

#fetching the calnder of a student
def fetch_student_calender():

    global caltk
    caltk = Tk()

    cal=pd.read_csv('Admin_data.csv')
    cal=cal[cal['A_Sheet']==s_1]

    sample=cal['A_Sheet']
    sample=str(sample)
    s1=""
    flag=0
    startind=0
    for i in range(0,len(sample)):
        if(flag==0):
            if((sample[i]>='0' and sample[i]<='9') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='a' and sample[i]<='z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='A' and sample[i]<='Z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
        else:
            break
    sample=sample[i-1:]
    for i in sample:
        if(i=="\n" or i==' '):
            break
        s1+=i
    s1+=".xlsx"

    li2=[]
    wb1=openpyxl.load_workbook(s1)
    name='data'
    f=0
    sheet=wb1[name]
    for row in sheet.iter_rows(min_row=2, min_col=2, max_row=350, max_col=2):
        for c in row:
            if(c.value==None):
                f=1
                break
            li2.append(c.value)
        if f==1:
            break
    wb1.save(s1)
    name=rno.get()
    name=name.upper()

    if name in li2:
        att_t.destroy()
        li=[]
        wb1=openpyxl.load_workbook(s1)
        f=0
        sheet=wb1[name]
        for row in sheet.iter_rows(min_row=2, min_col=1, max_row=350, max_col=1):
            for c in row:
                if(c.value==None):
                    f=1
                    break
                li.append(c.value)
            if f==1:
                break
        wb1.save(s1)

        li2=[]
        wb1=openpyxl.load_workbook(s1)
        name=rno.get()
        name=name.upper()
        f=0
        sheet=wb1[name]
        for row in sheet.iter_rows(min_row=2, min_col=2, max_row=350, max_col=2):
            for c in row:
                if(c.value==None):
                    
                    f=1
                    break
                li2.append(c.value)
            if f==1:
                break
        wb1.save(s1)

        cal_df=pd.DataFrame(columns=['Date','Present/Absent'])
        cal_df['Date']=li
        cal_df['Present/Absent']=li2

        caltk.geometry('1920x1080')

        txt=Text(caltk) 
        txt.pack() 

        class PrintToTXT(object): 
            def write(self, s): 
                txt.insert(END, s)

        sys.stdout = PrintToTXT() 
        print('Technical Hub Users and Passwords') 
        print(cal_df)

    else:
        messagebox.showinfo("Invalid","Please enter a valid Roll No!")
        
    Button(caltk,text="Back",height="1",width=20,bd=1,command=call_back).place(relx=0.5,rely=0.65,anchor=CENTER)
    Button(caltk,text="Return Trainer Home",height="1",width=20,bd=1,command=call_calth).place(relx=0.5,rely=0.7,anchor=CENTER)

    caltk.mainloop()

#caling trainer properties
def call_back():
    caltk.destroy()
    attendance_tracker()
    
def call_calth():
    caltk.destroy()
    trainer_properties(user_t,t_d,s_1)
    

#calling trainer login from trainer properties
def call_tlogin():
    tp.destroy()
    trainerlogintab()
    
#logintab() funtion  
def trainerlogintab():
    
    global root
    global username
    global password
    global id_
    
    root=Tk()
    root.title("Attendance Drive Ready+ (AWS DevOps)")
    root.geometry("1920x1080")
    root.configure(bg="#ffffff")
    bc1=Frame(root,bg="#fad311",width=650,height=450).place(relx=0.5,rely=0.45,anchor=CENTER)
    bc2=Frame(root,bg="#72bd20",width=600,height=400).place(relx=0.5,rely=0.45,anchor=CENTER)

    ltitle=Label(text="Hello Trainer!!",height=0,font=("Gabriola",50,'bold'),bg="#72bd20").place(relx=0.5,rely=0.32,anchor=CENTER)

    bg=PhotoImage(file = "logo_header.png")
    img=Label(root,image=bg)
    img.place(x=0,y=0)
    
    Label(root,text="ID: ",font=("Sitka Small Semibold",10),bg="#72bd20").place(relx=0.48,rely=0.4,anchor=CENTER)
    Label(root,text="Username: ",font=("Sitka Small Semibold",10),bg="#72bd20").place(relx=0.45,rely=0.45,anchor=CENTER)
    Label(root,text="Password: ",font=("Sitka Small Semibold",10),bg="#72bd20").place(relx=0.45,rely=0.5,anchor=CENTER)

    username=StringVar()
    password=StringVar()
    id_=StringVar()

    i_d=Entry(root,textvariable=id_,width=22,bd=2,font=("Californian FB",10)).place(relx=0.55,rely=0.4,anchor=CENTER)
    eun=Entry(root,textvariable=username,width=22,bd=2,font=("Californian FB",10)).place(relx=0.55,rely=0.45,anchor=CENTER)
    eps=Entry(root,textvariable=password,width=22,bd=2,font=("Californian FB",10)).place(relx=0.55,rely=0.5,anchor=CENTER)

    Button(root,text="Home",height="1",width=10,bd=1,command=exitfun).place(relx=0.45,rely=0.62,anchor=CENTER)
    Button(root,text="Login",height="1",width=10,bd=1,command=login).place(relx=0.55,rely=0.62,anchor=CENTER)           
    
    root.mainloop()

#exit()
def exitfun():
    root.destroy()
    homepage()

###################################################################################################

                                    #TRAINER END
    
####################################################################################################

#***************************************************************************************************************************
#***************************************************************************************************************************

###################################################################################################

                                    #ADMIN
    
####################################################################################################
##########################################################  ///viewing User
#view Users for Admin
def view_users():
    global ad
    
    ad = Tk()
    ad.geometry('1920x1080')
    
    dframe=pd.read_csv("Admin_data.csv",index_col=[0])

    txt=Text(ad) 
    txt.pack() 

    class PrintToTXT(object): 
        def write(self, s): 
            txt.insert(END, s)

    sys.stdout = PrintToTXT() 
    print('Technical Hub Users and Passwords') 
    print(dframe)

    Button(ad,text="Return Admin Home",height="1",width=20,bd=1,command=call_adminv).place(relx=0.5,rely=0.7,anchor=CENTER)

    ad.mainloop()

#calling viewer
def call_viewers():
    admin.destroy()
    view_users()

#calling admin from view
def call_adminv():
    ad.destroy()
    Admin()
##########################################################  ///viewing User
    
##########################################################  ///adiing User
#calling add Event
def call_addevent():
    admin.destroy()
    add_event()

# addevent for a Admin
def add_event():
    global a_eventtk
    global a_evid
    global a_evus
    global a_evps
    global a_ev
    
    global evaid
    global evus
    global evps
    global ev
    
    a_eventtk=Tk()
    a_eventtk.geometry("1920x1080")
    a_eventtk.configure(bg="#ffffff")
    
    bc1=Frame(a_eventtk,bg="#fad311",width=650,height=450).place(relx=0.5,rely=0.45,anchor=CENTER)
    bc2=Frame(a_eventtk,bg="#72bd20",width=600,height=400).place(relx=0.5,rely=0.45,anchor=CENTER)

    Label(text="Please enter new Username and Password",height=0,font=("Gabriola",20,'bold'),bg="#72bd20").place(relx=0.5,rely=0.3,anchor=CENTER)

    Label(a_eventtk,text="ID: ",font=("Sitka Small Semibold",12),bg="#72bd20").place(relx=0.48,rely=0.35,anchor=CENTER)
    Label(a_eventtk,text="Username: ",font=("Sitka Small Semibold",12),bg="#72bd20").place(relx=0.45,rely=0.4,anchor=CENTER)
    Label(a_eventtk,text="Password: ",font=("Sitka Small Semibold",12),bg="#72bd20").place(relx=0.45,rely=0.45,anchor=CENTER)
    Label(a_eventtk,text="Domain: ",font=("Sitka Small Semibold",12),bg="#72bd20").place(relx=0.45,rely=0.5,anchor=CENTER)
    
    a_evid=StringVar()
    a_evus=StringVar()
    a_evps=StringVar()
    a_ev=StringVar()

    evaid=Entry(a_eventtk,textvariable=a_evid,width=22,bd=2,font=("Californian FB",10))
    evaid.place(relx=0.55,rely=0.35,anchor=CENTER)
    evus=Entry(a_eventtk,textvariable=a_evus,width=22,bd=2,font=("Californian FB",10))
    evus.place(relx=0.55,rely=0.4,anchor=CENTER)
    evps=Entry(a_eventtk,textvariable=a_evps,width=22,bd=2,font=("Californian FB",10))
    evps.place(relx=0.55,rely=0.45,anchor=CENTER)
    ev=Entry(a_eventtk,textvariable=a_ev,width=22,bd=2,font=("Californian FB",10))
    ev.place(relx=0.55,rely=0.5,anchor=CENTER)
    
    Button(a_eventtk,text="Done",height="1",width=10,bd=1,command=add_e).place(relx=0.5,rely=0.6,anchor=CENTER)
    Button(a_eventtk,text="Return Admin Home",height="1",width=20,bd=1,command=close_adevent).place(relx=0.5,rely=0.7,anchor=CENTER)

    a_eventtk.mainloop()

#adding to Admin data
def add_e():

    did=a_evid.get()
    d1=a_evus.get()
    d2=a_evps.get()
    d3=a_ev.get()
    d4=did+"_sheet"

    hsdf=pd.read_csv('Admin_data.csv')

    if(did =='' or d1=='' or d2=='' or d3==''):
        messagebox.showinfo("Invalid","password can't be empty!")

    elif ( int(did) in hsdf['ID'].unique() ):
        messagebox.showinfo("Invalid","ID already exists.")
    else:
        
        df1=pd.read_csv('Admin_data.csv')
    
        df=pd.DataFrame(columns=['ID','username','password','domain','A_Sheet'])
        new_row=pd.DataFrame({'ID':did,'username':d1,'password':d2,'domain':d3,'A_Sheet':d4}, index=[0])
        df=pd.concat([new_row,df.loc[:]]).reset_index(drop=True)
        df.set_index("ID", inplace = True)
        d=df1.shape[0]

        Label(a_eventtk,text=d1+" event added",font=("Comic Sans Ms",12),bg="#72bd20").place(relx=0.5,rely=0.55,anchor=CENTER)

        df.to_csv('Admin_data.csv', mode='a', header=False)

        tu_df=pd.read_csv('Admin_data.csv')
        tu_df=tu_df[tu_df['A_Sheet']==d4]

        s1=''
        s1+=d4+".xlsx"
        a_edata=pd.DataFrame(columns=['Roll_No'])

        wb=openpyxl.Workbook()
        wb.save(s1)
        
        with pd.ExcelWriter(s1, mode="a", engine="openpyxl",if_sheet_exists='overlay') as writer:
            a_edata.to_excel(writer, sheet_name='data')

        workbook=openpyxl.load_workbook(s1)
        del workbook['Sheet']
        workbook.save(s1)
            
    clear_adevent()

#clearig add_user
def clear_adevent():
    evus.delete(0,END)
    evps.delete(0,END)
    evaid.delete(0,END)
    ev.delete(0,END)

def close_adevent():
    a_eventtk.destroy()
    Admin()
    
#calling add user
def call_adduser():
    admin.destroy()
    add_user()
    
#Add a user for admin
def add_user():

    global a_usertk
    global a_id
    global a_us
    global a_ps
    global a_do
    
    global aid
    global us
    global ps
    global do
    
    a_usertk=Tk()
    a_usertk.geometry("1920x1080")
    a_usertk.configure(bg="#ffffff")
    
    bc1=Frame(a_usertk,bg="#fad311",width=650,height=450).place(relx=0.5,rely=0.45,anchor=CENTER)
    bc2=Frame(a_usertk,bg="#72bd20",width=600,height=400).place(relx=0.5,rely=0.45,anchor=CENTER)

    Label(text="Please enter new Username and Password",height=0,font=("Gabriola",20,'bold'),bg="#72bd20").place(relx=0.5,rely=0.3,anchor=CENTER)

    Label(a_usertk,text="ID: ",font=("Sitka Small Semibold",12),bg="#72bd20").place(relx=0.48,rely=0.35,anchor=CENTER)
    Label(a_usertk,text="Username: ",font=("Sitka Small Semibold",12),bg="#72bd20").place(relx=0.45,rely=0.4,anchor=CENTER)
    Label(a_usertk,text="Password: ",font=("Sitka Small Semibold",12),bg="#72bd20").place(relx=0.45,rely=0.45,anchor=CENTER)
    Label(a_usertk,text="Domain: ",font=("Sitka Small Semibold",12),bg="#72bd20").place(relx=0.45,rely=0.5,anchor=CENTER)
    
    a_id=StringVar()
    a_us=StringVar()
    a_ps=StringVar()
    a_do=StringVar()

    aid=Entry(a_usertk,textvariable=a_id,width=22,bd=2,font=("Californian FB",10))
    aid.place(relx=0.55,rely=0.35,anchor=CENTER)
    us=Entry(a_usertk,textvariable=a_us,width=22,bd=2,font=("Californian FB",10))
    us.place(relx=0.55,rely=0.4,anchor=CENTER)
    ps=Entry(a_usertk,textvariable=a_ps,width=22,bd=2,font=("Californian FB",10))
    ps.place(relx=0.55,rely=0.45,anchor=CENTER)
    do=Entry(a_usertk,textvariable=a_do,width=22,bd=2,font=("Californian FB",10))
    do.place(relx=0.55,rely=0.5,anchor=CENTER)
    
    Button(a_usertk,text="Done",height="1",width=10,bd=1,command=add_d).place(relx=0.5,rely=0.6,anchor=CENTER)
    Button(a_usertk,text="Return Admin Home",height="1",width=20,bd=1,command=close_aduser).place(relx=0.5,rely=0.7,anchor=CENTER)

    a_usertk.mainloop()
    
#closing add_user
def close_aduser():
    a_usertk.destroy()
    Admin()
    
#adding to Admin data
def add_d():

    did=a_id.get()
    d1=a_us.get()
    d2=a_ps.get()
    d3=a_do.get()
    d4=did+"_sheet"

    hsdf=pd.read_csv('Admin_data.csv')

    if(did =='' or d1=='' or d2=='' or d3==''):
        messagebox.showinfo("Invalid","password can't be empty!")

    elif ( int(did) in hsdf['ID'].unique() ):
        messagebox.showinfo("Invalid","ID already exists.")
    else:
        
        df1=pd.read_csv('Admin_data.csv')
    
        df=pd.DataFrame(columns=['ID','username','password','domain','A_Sheet'])
        new_row=pd.DataFrame({'ID':did,'username':d1,'password':d2,'domain':d3,'A_Sheet':d4}, index=[0])
        df=pd.concat([new_row,df.loc[:]]).reset_index(drop=True)
        df.set_index("ID", inplace = True)
        d=df1.shape[0]

        Label(a_usertk,text=d1+" added",font=("Comic Sans Ms",12),bg="#72bd20").place(relx=0.5,rely=0.55,anchor=CENTER)

        df.to_csv('Admin_data.csv', mode='a', header=False)

        tu_df=pd.read_csv('Admin_data.csv')
        tu_df=tu_df[tu_df['A_Sheet']==d4]

        s1=''
        s1+=d4+".xlsx"
        a_data=pd.DataFrame(columns=['Index','Roll_No','Email'])
        a_data.set_index("Index", inplace = True)

        wb=openpyxl.Workbook()
        wb.save(s1)
        
        with pd.ExcelWriter(s1, mode="a", engine="openpyxl",if_sheet_exists='overlay') as writer:
            a_data.to_excel(writer, sheet_name='data')

        workbook=openpyxl.load_workbook(s1)
        del workbook['Sheet']
        workbook.save(s1)
            
    clear_aduser()

#clearig add_user
def clear_aduser():
    us.delete(0,END)
    ps.delete(0,END)
    aid.delete(0,END)
    do.delete(0,END)
    
##########################################################  ///adiing User
    
##########################################################   ///DELETING
#calling deleting fun
def call_delete():
    admin.destroy()
    delete_user()
    
#deleting user data for admin
def delete_user():
    global delete_u
    global ind
    global ind_

    delete_u=Tk()
    delete_u.geometry("1920x1080")
    delete_u.configure(bg="#ffffff")

    bc1=Frame(delete_u,bg="#fad311",width=650,height=450).place(relx=0.5,rely=0.45,anchor=CENTER)
    bc2=Frame(delete_u,bg="#72bd20",width=600,height=400).place(relx=0.5,rely=0.45,anchor=CENTER)

    ind=StringVar()

    Label(delete_u,text="Please Enter ID No of Trainer Username:",font=("Sitka Small Semibold",15),bg="#72bd20").place(relx=0.5,rely=0.32,anchor=CENTER)
    
    ind_=Entry(delete_u,textvariable=ind,width=22,bd=2,font=("Californian FB",10))
    ind_.place(relx=0.5,rely=0.42,anchor=CENTER)

    Button(delete_u,text="Delete",height="1",width=10,bd=1,command=deleting).place(relx=0.5,rely=0.55,anchor=CENTER)

    Button(delete_u,text="Return to Admin page",height="1",width=20,bd=1,command=call_admind).place(relx=0.5,rely=0.7,anchor=CENTER)

    delete_u.mainloop()
    
#deleting user name
def deleting():
    x=ind.get()
    if x!='':
        
        x=int(x)
        dfa=pd.read_csv('Admin_data.csv')
        dfa.set_index("ID", inplace = True)
        dfa=dfa.drop([x])
        dfa.to_csv('Admin_data.csv')

        Label(delete_u,text="Username with ID: "+str(x)+" deleted",font=("Comic Sans Ms",10),bg="#72bd20").place(relx=0.5,rely=0.47,anchor=CENTER)
    else:
        messagebox.showinfo("Invalid","Input can't be empty!")
    ind_.delete(0,END)

#calling admin
def call_admind():
    delete_u.destroy()
    Admin()
########################################################## ///Deleting

########################################################## /// changing password
#changing password of admin
def call_changepass():
    admin.destroy()
    changepass()

#change password fun
def changepass():
    global cpass_a
    global c_pass1
    global c_pass2
    global a_cpass1
    global a_cpass2

    cpass_a=Tk()
    cpass_a.geometry("1920x1080")
    cpass_a.configure(bg="#ffffff")
    
    bc1=Frame(cpass_a,bg="#fad311",width=650,height=450).place(relx=0.5,rely=0.45,anchor=CENTER)
    bc2=Frame(cpass_a,bg="#72bd20",width=600,height=400).place(relx=0.5,rely=0.45,anchor=CENTER)

    Label(cpass_a,text="Please enter new Password",height=0,font=("Gabriola",20,'bold'),bg="#72bd20").place(relx=0.5,rely=0.3,anchor=CENTER)

    Label(cpass_a,text="Password: ",font=("Sitka Small Semibold",12),bg="#72bd20").place(relx=0.46,rely=0.42,anchor=CENTER)
    Label(cpass_a,text="Re-enter Password: ",font=("Sitka Small Semibold",12),bg="#72bd20").place(relx=0.43,rely=0.5,anchor=CENTER)

    c_pass1=StringVar()
    c_pass2=StringVar()
    
    a_cpass1=Entry(cpass_a,textvariable=c_pass1,width=22,bd=2,font=("Californian FB",10))
    a_cpass1.place(relx=0.55,rely=0.42,anchor=CENTER)
    a_cpass2=Entry(cpass_a,textvariable=c_pass2,width=22,bd=2,font=("Californian FB",10))
    a_cpass2.place(relx=0.55,rely=0.5,anchor=CENTER)

    img=PhotoImage(file="logo_header.png")
    img_label=Label(cpass_a,image=img)
    img_label.grid(row=0,column=0)
    img_label.image=img

    Button(cpass_a,text="change",height="1",width=10,bd=1,command=call_change).place(relx=0.5,rely=0.6,anchor=CENTER)
    Button(cpass_a,text="Return Home screen",height="1",width=20,bd=1,command=call_adminc).place(relx=0.5,rely=0.7,anchor=CENTER)

    cpass_a.mainloop()
    
#verifing and changing the passwords
def call_change():
    p1=c_pass1.get()
    p2=c_pass2.get()


    if p1=="" or p2=="":
        messagebox.showinfo("Invalid","password can't be empty!")
        a_cpass1.delete(0,END)
        a_cpass2.delete(0,END)
        
    elif(p1 == p2):
        dfa=pd.read_csv('Admin_data.csv')
        dfa.set_index("ID", inplace=True)
        idx=dfa.index[dfa['username']=='ADMIN']
        dfa.loc[idx,'password']=p1
        dfa.to_csv('Admin_data.csv')
        
        lab1=Label(cpass_a,text="Password changed",font=("Comic Sans MS",12),bg="#72bd20",fg="#0e6473")
        lab1.place(relx=0.5,rely=0.55,anchor=CENTER)
        a_cpass1.delete(0,END)
        a_cpass2.delete(0,END)
        
    elif p1!=p2:
        messagebox.showinfo("Invalid","passwords didn't matched!")
        a_cpass1.delete(0,END)
        a_cpass2.delete(0,END)

#calling admin from change password
def call_adminc():
    cpass_a.destroy()
    Admin()    
########################################################## /// changing password
#admin login tab
def Adminlogintab():
    
    global adlogin
    global a_username
    global a_password
    
    adlogin=Tk()
    adlogin.title("Attendance Drive Ready+ (AWS DevOps)")
    adlogin.geometry("1920x1080")
    adlogin.configure(bg="#ffffff")

    bc1=Frame(adlogin,bg="#fad311",width=650,height=450).place(relx=0.5,rely=0.45,anchor=CENTER)
    bc2=Frame(adlogin,bg="#72bd20",width=600,height=400).place(relx=0.5,rely=0.45,anchor=CENTER)

    
    bg=PhotoImage(file = "logo_header.png",master=adlogin)
    img=Label(adlogin,image=bg)
    img.place(x=0,y=0)
    
    ltitle=Label(text="Hello ADMIN!!",height=0,font=("Gabriola",50,'bold'),bg="#72bd20").place(relx=0.5,rely=0.32,anchor=CENTER)

    Label(adlogin,text="Username: ",font=("Sitka Small Semibold",15),bg="#72bd20").place(relx=0.45,rely=0.42,anchor=CENTER)
    Label(adlogin,text="Password: ",font=("Sitka Small Semibold",15),bg="#72bd20").place(relx=0.45,rely=0.5,anchor=CENTER)

    a_username=StringVar()
    a_password=StringVar() 

    a_eun=Entry(adlogin,textvariable=a_username,width=22,bd=2,font=("Californian FB",10)).place(relx=0.55,rely=0.42,anchor=CENTER)
    a_eps=Entry(adlogin,textvariable=a_password,width=22,bd=2,font=("Californian FB",10)).place(relx=0.55,rely=0.5,anchor=CENTER)

    Button(adlogin,text="Exit",height="1",width=10,bd=1,command=call_adminh).place(relx=0.45,rely=0.6,anchor=CENTER)
    Button(adlogin,text="Login",height="1",width=10,bd=1,command=admin_login).place(relx=0.55,rely=0.6,anchor=CENTER)           
    
    adlogin.mainloop()

#calling home page from admin
def call_adminh():
    
    adlogin.destroy()
    homepage()
    
#Admin login access
def admin_login():
    user=a_username.get()
    pas=a_password.get()
    ap=pd.read_csv('Admin_data.csv')
    ap=ap[ap['username']=='ADMIN']
    ap=ap['password']
    sample=str(ap)
    s=""
    flag=0
    startind=0
    for i in range(0,len(sample)):
        if(flag==0):
            if((sample[i]>='0' and sample[i]<='9') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='a' and sample[i]<='z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
                
            elif((sample[i]>='A' and sample[i]<='Z') and (sample[i+1]!=' ')):
                startind=i
                flag=1
        else:
            break
    sample=sample[i-1:]
    for i in sample:
        if(i=="\n" or i==' '):
            break
        s+=i
    
    if(user=="ADMIN" and pas==s):
        adlogin.destroy()
        Admin()
        
    elif user=="" or pas=="":
        messagebox.showinfo("Invalid","Username and Password can't be empty!")
    elif user!="ADMIN":
        messagebox.showinfo("Invalid","Please enter a valid username!")
    elif pas!=s:
        messagebox.showinfo("Username Error","Please enter the correct password!")
    else:
        messagebox.showinfo("Password Error","Please enter correct details!")


#Admin Properties
def Admin():

    global admin
    
    admin=Tk()
    admin.title("Technical Attendance Portal")
    admin.geometry("1920x1080")
    admin.configure(bg="#ffffff")

    bc1=Frame(admin,bg="#fad311",width=650,height=450).place(relx=0.5,rely=0.45,anchor=CENTER)
    bc2=Frame(admin,bg="#72bd20",width=600,height=400).place(relx=0.5,rely=0.45,anchor=CENTER)

    ltitle=Label(text="Hello Admin!!",height=0,font=("Gabriola",50,'bold'),bg="#72bd20").place(relx=0.5,rely=0.32,anchor=CENTER)

    img = PhotoImage(file="logo_header.png",master=admin)
    label = Label(admin,image=img)
    label.place(x=0,y=0)
    
    Button(admin,text="Add User",height="1",width=15,bd=1,command=call_adduser).place(relx=0.5,rely=0.4,anchor=CENTER)
    Button(admin,text="View Users",height="1",width=15,bd=1,command=call_viewers).place(relx=0.5,rely=0.45,anchor=CENTER)
    Button(admin,text="Delete User",height="1",width=15,bd=1,command=call_delete).place(relx=0.5,rely=0.5,anchor=CENTER)
    Button(admin,text="Change Password",height="1",width=15,bd=1,command=call_changepass).place(relx=0.5,rely=0.55,anchor=CENTER)
    Button(admin,text="Send Mails",height="1",width=15,bd=1,command=call_sendmails).place(relx=0.5,rely=0.6,anchor=CENTER)
    Button(admin,text="Add Event",height="1",width=15,bd=1,command=call_addevent).place(relx=0.5,rely=0.65,anchor=CENTER)
    Button(admin,text="Return Home",height="1",width=15,bd=1,command=call_homea).place(relx=0.5,rely=0.7,anchor=CENTER)
    
    admin.mainloop()

#calling hom
def call_homea():
    admin.destroy()
    homepage()

#calling sendmails
def call_sendmails():
    admin.destroy()
    
    global abc
    
    abc=Tk()
    abc.title("Technical Attendance Portal")
    abc.geometry("1920x1080")
    abc.configure(bg="#ffffff")

    bc1=Frame(abc,bg="#fad311",width=650,height=450).place(relx=0.5,rely=0.45,anchor=CENTER)
    bc2=Frame(abc,bg="#72bd20",width=600,height=400).place(relx=0.5,rely=0.45,anchor=CENTER)

    img = PhotoImage(file="logo_header.png",master=abc)
    label = Label(abc,image=img)
    label.place(x=0,y=0)

    Label(abc,text="Please press the button to send.",font=("Comic Sans MS",20),bg="#72bd20").place(relx=0.5,rely=0.3,anchor=CENTER)

    Button(abc,text="send",height="2",width=15,bd=1,command=sendmail_to_college).place(relx=0.5,rely=0.5,anchor=CENTER)

    abc.mainloop()
    
#sending mail to college function
def sendmail_to_college():
    
    smtc_df=pd.read_csv('Admin_data.csv')
    li=smtc_df['A_Sheet'].tolist()
    li.pop(0)
    college1=pd.DataFrame(columns=['Roll_No'])
    college2=pd.DataFrame(columns=['Roll_No'])
    college3=pd.DataFrame(columns=['Roll_No'])
    s1=''
    for i in li:
        s1=''
        s1+=i+".xlsx"
        wb1=openpyxl.load_workbook(s1)
        name='Attendance'
        sheet=wb1[name]
        r=0
        f=0 
        for row in sheet.iter_rows(min_row=2, min_col=2, max_row=350, max_col=2):
            for c in row:
                r+=1
                if(c.value==None):
                    f=1

                    break
                if 'A9' in c.value:
                    new_row=pd.DataFrame({'Roll_No':c.value}, index=[0])
                    college1 = pd.concat([new_row,df.loc[:]]).reset_index(drop=True)

                elif 'P3' in c.value:
                    new_row=pd.DataFrame({'Roll_No':c.value}, index=[0])
                    college2 = pd.concat([new_row,df.loc[:]]).reset_index(drop=True)

                elif 'MH' in c.value:
                    new_row=pd.DataFrame({'Roll_No':c.value}, index=[0])
                    college3 = pd.concat([new_row,df.loc[:]]).reset_index(drop=True)
            if f==1:
                break
        wb1.save(s1)
    college1.to_csv('AEC.csv')
    college2.to_csv('ACET.csv')
    college3.to_csv('ACOE.csv')
    sendemailclg('sanjusiddu1951@gmail.com','AEC.csv')
    sendemailclg('sanjusiddu1951@gmail.com','ACET.csv')
    sendemailclg('sanjusiddu1951@gmail.com','ACOE.csv')

    Label(abc,text="Sent.",font=("Comic Sans MS",30),bg="#72bd20").place(relx=0.5,rely=0.5,anchor=CENTER)

    Button(abc,text="Back",height="1",width=15,bd=1,command=call_adminmail).place(relx=0.5,rely=0.6,anchor=CENTER)
    abc.mainloop()

def call_adminmail():
    abc.destroy()
    Admin()

####################################################################################################
    
                                    #ADMIN END

####################################################################################################
    
#################################    Calling ADMIN and Trainer   #####################################

#calling logintab from homepage()
def call_trainer():
    home.destroy()
    trainerlogintab()

#calling admin
def call_admin():
    home.destroy()
    Adminlogintab()

#################################    Calling ADMIN and Trainer   #####################################
#Homepage
def homepage():

    global home
    
    home=Tk()
    home.title("Technical Attendance Portal")
    home.geometry("1920x1080")
    home.configure(bg="#ffffff")

    bc1=Frame(home,bg="#fad311",width=650,height=450).place(relx=0.5,rely=0.45,anchor=CENTER)
    bc2=Frame(home,bg="#72bd20",width=600,height=400).place(relx=0.5,rely=0.45,anchor=CENTER)

    ltitle=Label(text="Good Day!!",height=0,font=("Gabriola",50,'bold'),bg="#72bd20").place(relx=0.5,rely=0.35,anchor=CENTER)

    img = PhotoImage(file="logo_header.png",master=home)
    label = Label(home,image=img)
    label.place(x=0,y=0)

    Button(home,text="ADMIN",height="3",width=10,bd=1,command=call_admin).place(relx=0.45,rely=0.5,anchor=CENTER)
    Button(home,text="TRAINER",height="3",width=10,bd=1,command=call_trainer).place(relx=0.55,rely=0.5,anchor=CENTER)           

    home.mainloop()
    
#############################   MAIN()  #############################################
homepage()
